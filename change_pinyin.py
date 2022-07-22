import pinyin
import openpyxl
if __name__ == '__main__':
    excel_file = '/Users/carlguo/Desktop/PPMT Team.xlsx'
    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb["Teams"]
    list = []
    n = worksheet.max_row
    m = worksheet.max_column
    for i in range(2, n + 1):
        pk = worksheet.cell(i, 1).value
        if pk is None:
            continue
        pk = int(pk)
        team_name = worksheet.cell(i, 2).value
        division = worksheet.cell(i, 3).value
        school = worksheet.cell(i, 4).value
        j = 5
        team_roster = []
        while j <= m and worksheet.cell(i, j).value != None and worksheet.cell(i, j).value != '':
            if len(worksheet.cell(i, j).value) <= 3:
                chinese = worksheet.cell(i, j).value
                name_pinyin_arr = pinyin.get(chinese, delimiter=' ', format='strip').split(' ')
                name_pinyin = name_pinyin_arr[1][0].upper() + name_pinyin_arr[1][1:] + ''.join(name_pinyin_arr[2:])
                name_pinyin += ' ' + name_pinyin_arr[0][0].upper() + name_pinyin_arr[0][1:]
                worksheet.cell(row=i,column=j).value = name_pinyin

            team_roster.append(worksheet.cell(i, j).value)
            j += 1

        wb.save("testexcel.xlsx")