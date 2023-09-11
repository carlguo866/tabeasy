import random
import re
import string
import openpyxl
from django.contrib.auth.decorators import user_passes_test, login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.core.exceptions import ValidationError
from django.db import transaction
from django.forms import inlineformset_factory
from django.http import HttpResponseForbidden, HttpResponseNotFound, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.views.generic import UpdateView
from tabeasy.settings import DEBUG
from django.core.files.storage import FileSystemStorage

from accounts.models import User
from submission.forms import CharacterPronounsForm
from submission.models.paradigm import Paradigm, ParadigmPreferenceItem, ParadigmPreference, \
    experience_description_choices
from submission.models.section import Section, SubSection
from submission.models.spirit import Spirit
from tabeasy.settings import DEBUG
from tabeasy.utils.mixins import JudgeOnlyMixin, PassRequestToFormViewMixin, TabOnlyMixin
from tabeasy_secrets.secret import  str_int
from tourney.forms import RoundForm, UpdateConflictForm, UpdateJudgeFriendForm, PairingFormSet, PairingSubmitForm, \
    JudgeForm, CheckinJudgeForm, CompetitorPronounsForm, TournamentForm, CompetitorForm, TeamForm
from submission.models.ballot import Ballot
from submission.models.character import Character, CharacterPronouns
from tourney.models import Tournament
from tourney.models.judge import Judge
from tourney.models.round import Round, Pairing
from tourney.models.team import Team
from tourney.models.competitor import Competitor

def sort_teams(teams):
    return list(reversed(sorted(teams,
                        key=lambda x: (x.total_ballots, x.total_cs, x.total_pd))))

def index(request):
    return render(request, 'index.html')


@user_passes_test(lambda u: u.is_staff)
def results(request):
    tournament = request.user.tournament
    if tournament.split_division:
        div1_teams = sort_teams([team for team in Team.objects.filter(division='Disney')])
        div2_teams = sort_teams([team for team in Team.objects.filter(division='Universal')])
        dict = {'teams_ranked': [div1_teams, div2_teams]}
    else:
        teams = sort_teams([team for team in Team.objects.filter(user__tournament=tournament)])
        dict = {'teams_ranked': teams}
    return render(request, 'tourney/tab/results.html', dict)

@user_passes_test(lambda u: u.is_staff)
def individual_awards(request):
    tournament = request.user.tournament
    competitor_scores = [(member, (member.p_att,member.d_att) , (member.p_wit,member.d_wit))
                         for member in Competitor.objects.filter(team__user__tournament=tournament)]
    atts_ranked = [(member,'P',att_score[0])
                   for member, att_score, wit_score in competitor_scores
                   if tournament.judges == 1 or att_score[0] >= 10]+ \
                  [(member, 'D', att_score[1])
                   for member, att_score, wit_score in competitor_scores
                   if tournament.judges == 1 or att_score[1] >= 10]
    atts_ranked = sorted(atts_ranked, key=lambda x: -x[2])
    wits_ranked = [(member,'P',wit_score[0])
                   for member, att_score, wit_score in competitor_scores
                   if tournament.judges == 1 or wit_score[0] >= 10]+ \
                  [(member, 'D', wit_score[1])
                   for member, att_score, wit_score in competitor_scores
                   if tournament.judges == 1 or wit_score[1] >= 10]
    wits_ranked = sorted(wits_ranked, key=lambda x: -x[2])


    dict = {'ranked': zip(atts_ranked,wits_ranked)}
    return render(request, 'tourney/tab/individual_awards.html', dict)


@user_passes_test(lambda u: u.is_staff)
def next_pairing(request, round_num):
    tournament = request.user.tournament
    # if Pairing.objects.filter(tournament=tournament).exists():
    #     next_round = max([pairing.round_num for pairing in Pairing.objects.filter(tournament=tournament)])+1
    # else:
    #     next_round = 1
    next_round = round_num +1
    if tournament.split_division:
        if next_round % 2 == 0:
            i_d_teams = sort_teams([team for team in Team.objects.filter(division='Disney')
                                       if team.next_side(next_round) == 'd'])
            div1_p_teams = sort_teams([team for team in Team.objects.filter(division='Disney')
                                       if team.next_side(next_round) == 'p'])
            div2_d_teams = sort_teams([team for team in Team.objects.filter(division='Universal')
                                       if team.next_side(next_round) == 'd'])
            div2_p_teams = sort_teams([team for team in Team.objects.filter(division='Universal')
                                       if team.next_side(next_round) == 'p'])
        else:
            div1_teams = sort_teams([team for team in Team.objects.filter(division='Disney')])
            div1_p_teams = []
            div1_d_teams = []
            for i in range(0, len(div1_teams), 2):
                if random.randint(0, 1):
                    div1_p_teams.append(div1_teams[i])
                    div1_d_teams.append(div1_teams[i+1])
                else:
                    div1_p_teams.append(div1_teams[i+1])
                    div1_d_teams.append(div1_teams[i])
            div2_teams = sort_teams([team for team in Team.objects.filter(division='Universal')])
            div2_p_teams = []
            div2_d_teams = []
            for i in range(0, len(div2_teams), 2):
                if random.randint(0, 1):
                    div2_p_teams.append(div2_teams[i])
                    div2_d_teams.append(div2_teams[i + 1])
                else:
                    div2_p_teams.append(div2_teams[i + 1])
                    div2_d_teams.append(div2_teams[i])

        dict = {'next_round': next_round,
                'divs': ['Disney','Universal'],
                'teams':[zip(div1_p_teams, div1_d_teams),
                         zip(div2_p_teams, div2_d_teams)]
                }
    else:
        if next_round % 2 == 0:
            d_teams = sort_teams([team for team in Team.objects.filter(user__tournament=tournament)
                                       if team.next_side(next_round) == 'd'])
            p_teams = sort_teams([team for team in Team.objects.filter(user__tournament=tournament)
                                       if team.next_side(next_round) == 'p'])
        else:
            teams = sort_teams([team for team in Team.objects.filter(user__tournament=tournament)])
            p_teams = []
            d_teams = []
            for i in range(0, len(teams), 2):
                if random.randint(0, 1):
                    p_teams.append(teams[i])
                    d_teams.append(teams[i + 1])
                else:
                    p_teams.append(teams[i + 1])
                    d_teams.append(teams[i])
        dict = {'next_round': next_round,
                'divs': ['Teams'],
                'teams': [zip(p_teams, d_teams)]
                }
    return render(request, 'tourney/pairing/next_pairing.html', dict)

@user_passes_test(lambda u: u.is_staff)
def pairing_index(request):
    tournament = request.user.tournament
    round_num_lists = sorted(Pairing.objects.filter(tournament=tournament).values_list('round_num',flat=True).distinct())
    pairings = []
    for round_num in round_num_lists:
        pairings.append(Pairing.objects.filter(tournament=tournament,
                                               round_num=round_num).order_by('division'))
    if Pairing.objects.filter(tournament=tournament).exists():
        next_round = max([pairing.round_num for pairing in Pairing.objects.filter(tournament=tournament)]) + 1
    else:
        next_round = 1
    dict = {'pairings': pairings, 'next_round': next_round }
    if request.session.get('extra'):
        dict.update(request.session['extra'])
    return render(request, 'tourney/pairing/main.html', dict)

@user_passes_test(lambda u: u.is_staff)
def edit_pairing(request, round_num):
    tournament = request.user.tournament
    if DEBUG:
        RoundFormSet = inlineformset_factory(Pairing, Round, form=RoundForm, formset=PairingFormSet,
                                             max_num=int(tournament.division_team_num/2), validate_max=True,
                                             extra=int(tournament.division_team_num/2))
    else:
        RoundFormSet = inlineformset_factory(Pairing, Round, form=RoundForm, formset=PairingFormSet,
                                             max_num=int(tournament.division_team_num/2), validate_max=True,
                                             extra=int(tournament.division_team_num/2))

    if request.user.tournament.split_division:
        if not Pairing.objects.filter(round_num=round_num).exists():
            div1_pairing = Pairing.objects.create(round_num=round_num, division='Disney')
            div2_pairing = Pairing.objects.create(round_num=round_num, division='Universal')
        else:
            div1_pairing = Pairing.objects.filter(round_num=round_num).get(division='Disney')
            div2_pairing = Pairing.objects.filter(round_num=round_num).get(division='Universal')

        available_judges_pk = [judge.pk for judge in Judge.objects.all()
                               if judge.get_availability(div1_pairing.round_num)]
        judges = Judge.objects.filter(pk__in=available_judges_pk).order_by('-checkin','-preside', 'user__username').all()

        if request.method == "POST":
            div1_formset = RoundFormSet(request.POST, request.FILES, prefix='div1', instance=div1_pairing,
                                        form_kwargs={'pairing': div1_pairing, 'other_formset':None, 'request':request })
            div2_formset = RoundFormSet(request.POST, request.FILES, prefix='div2', instance=div2_pairing,
                                        form_kwargs={'pairing': div2_pairing, 'other_formset':div1_formset, 'request':request })

            div1_submit_form = PairingSubmitForm(request.POST, prefix='div1', instance=div1_pairing)
            div2_submit_form = PairingSubmitForm(request.POST, prefix='div2', instance=div2_pairing)

            if div1_submit_form.is_valid():
                div1_submit_form.save()
            if div2_submit_form.is_valid():
                div2_submit_form.save()
            both_true = True
            if div1_formset.is_valid():
                # get courtroom
                actual_round_num = len(div1_formset)
                for form in div1_formset:
                    if form.instance.p_team == None or form.instance.d_team == None:
                        actual_round_num -= 1
                if div1_formset[0].instance.pairing.division == 'Disney':
                    random_choice = string.ascii_uppercase[:8][:actual_round_num]
                else:
                    random_choice = string.ascii_uppercase[8:2 * 8][:actual_round_num]
                for round in Pairing.objects.get(pk=div1_pairing.pk).rounds.all():
                    if round.courtroom != None:
                        random_choice = random_choice.replace(round.courtroom, '')
                random_choice = random.sample(list(random_choice), len(random_choice))
                for form in div1_formset:
                    if form.instance.p_team != None and form.instance.d_team != None \
                            and form.instance.courtroom == None:
                        form.instance.courtroom = random_choice[0]
                        del(random_choice[0])
                        form.save()
                div1_formset.save()
            else:
                both_true = False

            if div2_formset.is_valid():
                actual_round_num = len(div2_formset)
                for form in div2_formset:
                    if form.instance.p_team == None or form.instance.d_team == None:
                        actual_round_num -= 1
                if div2_formset[0].instance.pairing.division == 'Disney':
                    random_choice = string.ascii_uppercase[:8][:actual_round_num]
                else:
                    random_choice = string.ascii_uppercase[8:2 * 8][:actual_round_num]
                for round in Pairing.objects.get(pk=div2_pairing.pk).rounds.all():
                    if round.courtroom != None:
                        random_choice = random_choice.replace(round.courtroom, '')
                random_choice = random.sample(list(random_choice), len(random_choice))
                for form in div2_formset:
                    if form.instance.p_team != None and form.instance.d_team != None \
                            and form.instance.courtroom == None:
                        form.instance.courtroom = random_choice[0]
                        del(random_choice[0])
                        form.save()
                div2_formset.save()
            else:
                both_true = False

            pairings = [div1_pairing, div2_pairing]
            for pairing in pairings:
                if pairing.final_submit and not pairing.publish:
                    for round in pairing.rounds.all():
                        if not Ballot.objects.filter(round=round).exists():
                            for judge in round.judges:
                                Ballot.objects.create(round=round, judge=judge)
                        else:
                            for judge in round.judges:
                                if not Ballot.objects.filter(round=round, judge=judge).exists():
                                    Ballot.objects.create(round=round, judge=judge)
                            for ballot in Ballot.objects.filter(round=round).all():
                                if ballot.judge not in round.judges:
                                    Ballot.objects.filter(round=round, judge=ballot.judge).delete()

            if both_true:
                return redirect('tourney:pairing_index')
        else:
            div1_formset = RoundFormSet(instance=div1_pairing,prefix='div1',
                                        form_kwargs={'pairing': div1_pairing,
                                                     'other_formset':None,
                                                     'request':request })
            div2_formset = RoundFormSet(instance=div2_pairing,prefix='div2',
                                        form_kwargs={'pairing': div2_pairing,
                                                     'other_formset':div1_formset,
                                                     'request':request })
            div1_submit_form = PairingSubmitForm(instance=div1_pairing, prefix='div1')
            div2_submit_form = PairingSubmitForm(instance=div2_pairing, prefix='div2')

        return render(request, 'tourney/pairing/edit.html', {'formsets': [div1_formset, div2_formset],
                                                             'submit_forms': [div1_submit_form, div2_submit_form],
                                                             'pairing': div1_pairing,
                                                             'judges': judges})
    else:
        if not Pairing.objects.filter(tournament=tournament, round_num=round_num).exists():
            pairing = Pairing.objects.create(tournament=tournament, round_num=round_num)
        else:
            pairing = Pairing.objects.get(tournament=tournament, round_num=round_num)

        available_judges_pk = [judge.pk for judge in Judge.objects.filter(user__tournament=tournament)
                               if judge.get_availability(pairing.round_num)]
        judges = Judge.objects.filter(pk__in=available_judges_pk).order_by('-checkin','-preside', 'user__username').all()

        if request.method == "POST":
            formset = RoundFormSet(request.POST, request.FILES, prefix='div1', instance=pairing,
                                        form_kwargs={'pairing': pairing,
                                                     'other_formset': None,
                                                     'request': request })
            submit_form = PairingSubmitForm(request.POST, prefix='div1', instance=pairing)

            if submit_form.is_valid():
                submit_form.save()

            both_true = True
            if formset.is_valid():
                # get courtroom
                actual_round_num = len(formset)
                for form in formset:
                    if form.instance.p_team == None or form.instance.d_team == None:
                        actual_round_num -= 1

                random_choice = string.ascii_uppercase[:int(tournament.division_team_num/2)][:actual_round_num]

                for round in Pairing.objects.get(pk=pairing.pk).rounds.all():
                    if round.courtroom != None:
                        random_choice = random_choice.replace(round.courtroom, '')
                random_choice = random.sample(list(random_choice), len(random_choice))
                for form in formset:
                    if form.instance.p_team != None and form.instance.d_team != None \
                            and form.instance.courtroom == None:
                        form.instance.courtroom = random_choice[0]
                        del (random_choice[0])
                        form.save()
                formset.save()
            else:
                both_true = False

            if both_true:
                pairings = [pairing]
                for pairing in pairings:
                    if pairing.final_submit and not pairing.publish:
                        for round in pairing.rounds.all():
                            if not Ballot.objects.filter(round=round).exists():
                                for judge in round.judges:
                                    Ballot.objects.create(round=round, judge=judge)
                            else:
                                for judge in round.judges:
                                    if not Ballot.objects.filter(round=round, judge=judge).exists():
                                        Ballot.objects.create(round=round, judge=judge)
                                for ballot in Ballot.objects.filter(round=round).all():
                                    if ballot.judge not in round.judges:
                                        Ballot.objects.filter(round=round, judge=ballot.judge).delete()
                return redirect('tourney:pairing_index')
        else:
            formset = RoundFormSet(instance=pairing, prefix='div1',
                                        form_kwargs={'pairing': pairing,
                                                     'other_formset': None,
                                                     'request': request})
            submit_form = PairingSubmitForm(instance=pairing, prefix='div1')

        return render(request, 'tourney/pairing/edit.html', {'formsets': [formset],
                                                             'submit_forms': [submit_form],
                                                             'pairing': pairing,
                                                             'judges': judges})

@user_passes_test(lambda u: u.is_staff)
def delete_pairing(request, round_num):
    errors = []
    cur_pairing = Pairing.objects.filter(tournament=request.user.tournament, round_num=round_num)
    if cur_pairing.exists():
        pairing_list = Pairing.objects.filter(tournament=request.user.tournament\
                                                     ).order_by('round_num')
        if pairing_list[len(pairing_list)-1] == cur_pairing[0]:
            Pairing.objects.filter(tournament=request.user.tournament, round_num=round_num).delete()
        else:
            errors.append('You can only delete the last pairing!')
    request.session['extra'] = {'errors': errors}
    return redirect('tourney:pairing_index')

        # request, 'tourney/pairing/main.html', {'errors':errors})

@login_required
def view_pairing(request, pk):
    pairing = Pairing.objects.get(pk=pk)
    if not pairing.team_submit:
        context = {}
    else:
        context = {'pairing': [pairing]}
    return render(request, 'tourney/pairing/view.html', context)

@user_passes_test(lambda u: u.is_staff)
def checkin_judges(request, round_num):
    if request.method == "POST":
        form = CheckinJudgeForm(request.POST, round_num=round_num, request=request)
        if form.is_valid():
            for judge in form.cleaned_data['checkins']:
                judge.checkin = True
                judge.save()
        return redirect('tourney:pairing_index')
    else:
        form = CheckinJudgeForm(round_num=round_num, request=request)

    return render(request, 'tourney/tab/checkin_judges.html',{'form':form, 'round_num':round_num})



@user_passes_test(lambda u: u.is_staff)
def view_teams(request):
    teams = Team.objects.filter(user__tournament=request.user.tournament)
    return render(request, 'tourney/tab/view_teams.html',{'teams': teams})



@user_passes_test(lambda u: u.is_staff)
def view_judges(request):
    judges = Judge.objects.filter(user__tournament=request.user.tournament)
    return render(request, 'tourney/tab/view_judges.html',{'judges': judges})

@user_passes_test(lambda u: u.is_staff)
def view_individual_judge(request, pk):
    judge = Judge.objects.get(pk=pk)

    if request.method == 'POST':
        user_form = UpdateConflictForm(data=request.POST, instance=judge, request=request)
        judge_form = JudgeForm(data=request.POST, instance=judge)
        if user_form.is_valid():
            user_form.save()
        if judge_form.is_valid():
            judge_form.save()
        return redirect('tourney:view_judges')
    else:
        user_form = UpdateConflictForm(instance=judge, request=request)
        judge_form = JudgeForm(instance=judge)

    context = {'conflict_form': user_form, 'preference_form': judge_form}
    return render(request, 'tourney/tab/view_individual_judge.html', context)


@login_required
def view_individual_team(request, pk):
    tournament = request.user.tournament
    # if not Team.objects.filter(user__tournament=tournament, pk=pk).exists():
    #     team = Team.objects.create(user__tournament=tournament)
    # else:
    team = Team.objects.get(user__tournament=tournament,pk=pk)
    if not (request.user.is_team and request.user.team == team) and not request.user.is_staff:
        return HttpResponseNotFound('<h1>Page not found</h1>')
    if team.byebuster:
        FormSet = inlineformset_factory(Team, Competitor, fields=('name', 'pronouns'),
                                        extra=6)
    else:
        FormSet = inlineformset_factory(Team, Competitor,fields=('name', 'pronouns'),
                                         max_num=12, validate_max=True,
                                         extra=6)

    if request.method == 'POST':
        formset = FormSet(request.POST, request.FILES,prefix='competitors', instance=team)
        team_form = TeamForm(data=request.POST, instance=team)
        if formset.is_valid() and team_form.is_valid():
            team_form.save()
            formset.save()
            return redirect('tourney:view_teams')
    else:
        formset = FormSet(prefix='competitors', instance=team)
        team_form = TeamForm(instance=team)

    context = {'formset': formset, 'team_form': team_form}
    return render(request, 'tourney/tab/view_individual_team.html', context)


@user_passes_test(lambda u: u.is_staff)
def edit_characters(request):
    tournament = request.user.tournament
    
    FormSet = inlineformset_factory(Tournament, Character,fields=('name', 'side'),
                                         max_num=12, validate_max=True,
                                         extra=6)

    if request.method == 'POST':
        formset = FormSet(request.POST, request.FILES,prefix='characters', instance=tournament)
        if formset.is_valid():
            formset.save()
            return redirect('index')
    else:
        formset = FormSet(prefix='characters', instance=tournament)

    context = {'formset': formset}
    return render(request, 'tourney/tab/edit_characters.html', context)



@user_passes_test(lambda u: u.is_staff)
def delete_individual_judge(request, pk):
    Judge.objects.get(pk=pk).delete()
    return redirect('tourney:view_judges')


@user_passes_test(lambda u: u.is_staff)
def delete_individual_team(request, pk):
    Team.objects.get(pk=pk).delete()
    return redirect('tourney:view_teams')



@user_passes_test(lambda u: u.is_staff)
def clear_checkin(request):
    Judge.objects.update(checkin=False)
    return redirect('tourney:pairing_index')

@user_passes_test(lambda u: u.is_staff)
def checkin_all_judges(request, round_num):
    available_judges_pk = [judge.pk for judge in Judge.objects.filter(user__tournament=request.user.tournament)
                           if judge.get_availability(round_num)]
    Judge.objects.filter(pk__in=available_judges_pk).update(checkin=True)
    return redirect('tourney:pairing_index')

@user_passes_test(lambda u: u.is_staff)
def view_ballot_status(request, pairing_id):
    pairing = Pairing.objects.get(pk=pairing_id)
    ballots = []
    for round in pairing.rounds.all():
        for ballot in round.ballots.all():
            ballots.append(ballot)
    ballots = sorted(ballots, key=lambda x: x.round.courtroom)
    return render(request, 'tourney/tab/view_ballots_status.html', {'ballots': ballots})


@user_passes_test(lambda u: u.is_staff)
def view_spirit_status(request): 
    tournament = request.user.tournament
    teams = Team.objects.filter(user__tournament=tournament)
    teams = sorted(teams, key=lambda x: x.spirit_score, reverse=True)
    return render(request, 'tourney/tab/view_spirit_status.html', {'teams': teams})

@user_passes_test(lambda u: u.is_staff)
def add_spirit_forms(request): 
    tournament = request.user.tournament
    teams = Team.objects.filter(user__tournament=tournament)
    for team in teams:
        if not Spirit.objects.filter(team=team).exists():
            spirit = Spirit.objects.create(team=team)
    return redirect('tourney:view_spirit_status')
        

@user_passes_test(lambda u: u.is_staff)
def view_captains_meeting_status(request, pairing_id):
    pairing = Pairing.objects.get(pk=pairing_id)
    captains_meetings = []
    for round in pairing.rounds.all():
        captains_meetings.append(round.captains_meeting)
    captains_meetings = sorted(captains_meetings, key=lambda x: x.round.courtroom)
    return render(request, 'tourney/tab/view_captains_meeting_status.html',
                  {'captains_meetings': captains_meetings})

# @login_required
# # def add_conflict(request):
# #     if request.method == 'POST':
# #         form = AddConflictForm(data=request.POST)
# #         if form.is_valid():
# #
# #     return render(request, 'tourney/add_conflict.html', {'form':form})

class TournamentUpdateView(TabOnlyMixin, UpdateView):
    model = Tournament
    form_class = TournamentForm
    template_name = 'utils/generic_form_help_text.html'

    def get_object(self, queryset=None):
        return self.request.user.tournament

    success_url = reverse_lazy('load_sections')



class ConflictUpdateView(JudgeOnlyMixin, PassRequestToFormViewMixin, UpdateView):
    model = Judge
    template_name = "tourney/add_conflict.html"

    form_class = UpdateConflictForm

    def get_form(self, form_class=None):
        form = super(ConflictUpdateView, self).get_form(form_class)
        form.fields['conflicts'].required = False
        return form

    def get_object(self, queryset=None):
        return self.request.user.judge

    success_url = reverse_lazy('index')

class JudgeFriendUpdateView(JudgeOnlyMixin, PassRequestToFormViewMixin, UpdateView):
    model = Judge
    template_name = "utils/generic_form.html"

    form_class = UpdateJudgeFriendForm

    def get_form(self, form_class=None):
        form = super(JudgeFriendUpdateView, self).get_form(form_class)
        form.fields['judge_friends'].required = False
        return form

    def get_object(self, queryset=None):
        return self.request.user.judge

    success_url = reverse_lazy('index')

class JudgePreferenceUpdateView(JudgeOnlyMixin, UpdateView):
    model = Judge
    template_name = "utils/generic_form.html"

    form_class = JudgeForm

    def get_object(self, queryset=None):
        return self.request.user.judge

    success_url = reverse_lazy('index')

@user_passes_test(lambda u: u.is_team)
def edit_competitor_pronouns(request):
    team = request.user.team
    if request.method == 'POST':
        competitor_pronouns_forms = [CompetitorPronounsForm(request.POST, instance=competitor,
                                                           prefix=competitor.name)
                                     for competitor in team.competitors.all()]
        for form in competitor_pronouns_forms:
            if form.is_valid():
                form.save()

        return redirect('index')
    else:
        competitor_pronouns_forms = [CompetitorPronounsForm(instance=competitor,
                                                           prefix=competitor.name)
                                     for competitor in team.competitors.all()]
    return render(request, 'tourney/competitor_pronouns.html', {
        'team': team,
        'forms': competitor_pronouns_forms
    })

@user_passes_test(lambda u: u.is_staff)
def generate_passwords(request): 
    if "GET" == request.method:
        return render(request, 'admin/load_excel.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Teams"]
        n = worksheet.max_row
        m = worksheet.max_column
        wb_changed = False
        for i in range(2, n + 1):
            if not worksheet.cell(row=i, column=17).value:
                wb_changed = True
                worksheet.cell(row=i, column=17).value = ''.join(
                random.choices(string.ascii_letters + string.digits, k=4))
            if not worksheet.cell(row=i, column=16).value:
                wb_changed = True
                worksheet.cell(row=i, column=16).value = request.user.tournament.short_name+'_'+''.join(
                    worksheet.cell(row=i, column=1).value.split(' '))
                
        worksheet = wb["Judges"]
        n = worksheet.max_row
        m = worksheet.max_column
        wb_changed = False
        for i in range(2, n + 1):
            first_name = worksheet.cell(i, 1).value
            last_name = worksheet.cell(i, 2).value
            
            if not worksheet.cell(row=i, column=10).value:
                wb_changed = True
                worksheet.cell(row=i, column=10).value = ''.join(
                    random.choices(string.ascii_letters + string.digits, k=4))
            if not worksheet.cell(row=i, column=9).value and first_name and last_name:
                wb_changed = True
                worksheet.cell(
                    row=i, column=9).value = f"{request.user.tournament.short_name}_{first_name.lower()}_{last_name.lower()}"
                
                
        response = HttpResponse(content_type='application/vnd.ms-excel')
        wb.save(response)
        return response

@user_passes_test(lambda u: u.is_staff)
def load_teams_and_judges(request): 
    if "GET" == request.method:
        return render(request, 'admin/load_excel.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        team_list, wb_changed = load_teams_wrapper(request, wb)
        judge_list, wb_changed2 = load_judges_wrapper(request, wb)
        return render(request, 'admin/load_excel.html', {"list": team_list + judge_list})


@user_passes_test(lambda u: u.is_staff)
def load_teams(request):
    if "GET" == request.method:
        return render(request, 'admin/load_excel.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        response_list, wb_changed = load_teams_wrapper(request, wb)
        response = HttpResponse(content_type='application/vnd.ms-excel')
        wb.save(response)
        return render(request, 'admin/load_excel.html', {"list": response_list})



def load_teams_wrapper(request, wb):
    worksheet = wb["Teams"]
    list = []
    n = worksheet.max_row
    m = worksheet.max_column
    wb_changed = False
    for i in range(2, n + 1):
        team_name = worksheet.cell(i, 1).value
        if Team.objects.filter(user__tournament=request.user.tournament, team_name=team_name).exists():
            pk = Team.objects.get(user__tournament=request.user.tournament, team_name=team_name).pk
        else:
            pk = None
            
        school = worksheet.cell(i, 2).value
        j = 3
        team_roster = []
        while j <= m and worksheet.cell(i,j).value != None and worksheet.cell(i,j).value != '':
            team_roster.append(worksheet.cell(i,j).value)
            j+=1
        message = ''
        username = worksheet.cell(i,16).value
        try:
            if Team.objects.filter(pk=pk).exists():
                Team.objects.filter(pk=pk).update(team_name=team_name,school=school)
                team = Team.objects.get(pk=pk)
                message += f' update {team_name} \n'
            else:
                raw_password = worksheet.cell(i,17).value
                if raw_password:
                    user = User(username=username, raw_password=raw_password, is_team=True, is_judge=False,
                                tournament=request.user.tournament)
                    user.set_password(raw_password)
                    user.save()
                    with transaction.atomic():
                        team = Team(user=user, team_name=team_name, school=school)
                        team.save()
                    message += f' create {team_name} \n'
            created_roster = [] 
            updated_roster = [] 
            for name in team_roster:
                name = re.sub(r'\([^)]*\)', '', name).strip()
                if Competitor.objects.filter(team=team, name=name).exists():
                    updated_roster.append(name)
                    Competitor.objects.filter(team=team, name=name).update(team=team,name=name)
                else:
                    created_roster.append(name)
                    Competitor.objects.create(name=name, team=team)
            if created_roster:
                str_created_roster = ' , '.join(created_roster)
                message += f' created roster {str_created_roster} \n'
            if updated_roster:
                str_updated_roster = ','.join(updated_roster)
                message += f' updated roster {str_updated_roster} \n'
            
        except Exception as e:
            message += str(e)
        else:
            message = ' SUCCESS ' + message
            
        list.append(message)
    return list, wb_changed

def load_judges_wrapper(request, wb):
    worksheet = wb["Judges"]
    list = []
    n = worksheet.max_row
    m = worksheet.max_column
    wb_changed = False
    for i in range(2, n + 1):
        first_name = worksheet.cell(i, 1).value
        last_name = worksheet.cell(i, 2).value
        
        if not worksheet.cell(row=i, column=10).value:
            wb_changed = True
            worksheet.cell(row=i, column=10).value = ''.join(
                random.choices(string.ascii_letters + string.digits, k=4))
        if not worksheet.cell(row=i, column=9).value and first_name and last_name:
            wb_changed = True
            worksheet.cell(
                row=i, column=9).value = f"{first_name.lower()}_{last_name.lower()}"
            
        username = worksheet.cell(i, 9).value
        if username == None or username == '':
            continue

        if last_name == None or last_name == '':
            last_name = ' '
        raw_password = worksheet.cell(i,10).value
        preside = worksheet.cell(i,3).value
        if preside in ['CIN','No preference']:
            preside = 2
        elif preside in ['Y', 'Presiding', 'Yes', 'y', 'YES']:
            preside = 1
        else:
            preside = 0
        availability = []
        for j in range(4, 9):
            if worksheet.cell(i, j).value in ['y', 'YES', 'Y', 'Yes']:
                availability.append(True)
            else:
                availability.append(False)
        
        message = ''
        try:
            if Judge.objects.filter(user__username=username).exists():
                message += f'update judge {username} \n'
                judge = Judge.objects.get(user__username=username)
                user = judge.user
                user.first_name = first_name
                user.last_name = last_name
                user.tournament = request.user.tournament
                user.save()


                judge.preside = preside
                for i in range(len(availability)):
                    setattr(judge, f'available_round{i+1}', availability[i])
                judge.save()
            else:
                message += f'create judge {username} \n'
                user = User(username=username,
                            first_name=first_name, last_name=last_name,
                            is_team=False, is_judge=True, tournament=request.user.tournament)
                user.set_password(raw_password)
                user.save()
                judge = Judge(user=user, preside=preside)
                for i in range(len(availability)):
                    setattr(judge, f'available_round{i+1}', availability[i])

                judge.save()

        except Exception as e:
            message += str(e)
        else:
            message = ' SUCCESS ' + message
        list.append(message)
    return list, wb_changed

@user_passes_test(lambda u: u.is_staff)
def load_judges(request):
    if "GET" == request.method:
        return render(request, 'admin/load_excel.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        response_list, wb_changed = load_judges_wrapper(request, wb)
        response = HttpResponse(content_type='application/vnd.ms-excel')
        wb.save(response)
        if wb_changed:
            return response
        elif DEBUG: 
            return render(request, 'admin/load_excel.html', {"list": response_list, })


@user_passes_test(lambda u: u.is_staff)
def load_paradigms(request):
    if "GET" == request.method:
        return render(request, 'admin/load_excel.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Paradigms"]
        list = []
        n = worksheet.max_row
        m = worksheet.max_column
        headers = [None]
        for i in range(1, m):
            headers.append(worksheet.cell(1, i).value)

        for i in range(2, n + 1):
            username = worksheet.cell(i, 1).value
            if username == None or username == '':
                continue
            paradigm_items = []
            for j in range(2, worksheet.max_column):
                value = worksheet.cell(i, j).value
                if value:
                    paradigm_items.append((headers[j], value))

            message = ''
            try:
                if not Judge.objects.filter(user__username=username).exists():
                    continue

                if Paradigm.objects.filter(judge__user__username=username).exists():
                    message += f'update judge paradigm {username}'
                    paradigm = Paradigm.objects.get(judge__user__username=username)
                else:
                    paradigm = Paradigm.objects.create(judge=Judge.objects.get(user__username=username))

                for name, value in paradigm_items:
                    if name == 'experience_description':
                        experiences = value.split(',')
                        experiences_actual_vals = []
                        for experience in experiences:
                            experience = experience.strip()
                            # message += str(experience)
                            for (actual_val, display_val) in experience_description_choices:
                                if experience == display_val:
                                    experiences_actual_vals.append(actual_val)
                        message += str(experiences_actual_vals)
                        setattr(paradigm, name, experiences_actual_vals)
                    elif name == 'experience_years':
                        setattr(paradigm, name, int(value))
                    else:
                        try:
                            paradigm_preference_pk = int(name)
                            if ParadigmPreferenceItem.objects.filter(
                                    paradigm=paradigm,paradigm_preference__pk=paradigm_preference_pk).exists():
                                ParadigmPreferenceItem.objects.filter(
                                    paradigm=paradigm, paradigm_preference__pk=paradigm_preference_pk).update(scale=int(value))
                            else:
                                ParadigmPreferenceItem.objects.create(
                                    paradigm=paradigm, paradigm_preference=ParadigmPreference.objects.get(pk=paradigm_preference_pk),
                                    scale=int(value))
                        except ValueError:
                            setattr(paradigm, name, value)
                paradigm.save()
            except Exception as e:
                message += str(e)
            else:
                message += ' success '
            list.append(message)
        return render(request, 'admin/load_excel.html', {"list": list})


amta_witnesses = {
        'P': ['Ari Felder', 'Aubrey Roy', 'Drew Hubbard', 'Jamie Savchenko'],
        'D': ['Casey Koller', 'Kennedy Heisman', 'R. Moore'],
        'other': ['D.B. Gelfand', 'Mandy Navarra', 'Shannon Shahid'],
    }


@user_passes_test(lambda u: u.is_staff)
def load_sections(request):
    tournament = request.user.tournament
    if not Section.objects.filter(tournament=tournament).exists():
        wit_nums = tournament.wit_nums
        i = 1
        openings = Section.objects.create(name='Openings', tournament=tournament)
        SubSection.objects.create(name=f'{tournament.p_choice} Opening',
                                  section=openings,
                                  side='P',
                                  role='att',
                                  type='statement',
                                  help_text=f'{tournament.p_choice} Opening',
                                  sequence=i)
        SubSection.objects.create(name=f'Defense Opening',
                                  section=openings,
                                  side='D',
                                  role='att',
                                  type='statement',
                                  help_text=f'Defense Opening',
                                  sequence=i)
        i+=1
        side_choices = {
            'P': tournament.p_choice,
            'D': 'Defense'
        }
        for side in ['P', 'D']:
            for wit_num in range(1, wit_nums+1):

                section = Section.objects.create(name=f'{side_choices[side]} Witness #{wit_num}', tournament=tournament)
                SubSection.objects.create(name=f'{side} Wit {wit_num} Wit Direct',
                                          section=section,
                                          side=side,
                                          role='wit',
                                          type='direct',
                                          help_text=f'Witness (Direct)',
                                          sequence=i)
                SubSection.objects.create(name=f'{side} Wit {wit_num} Att Direct',
                                          section=section,
                                          side=side,
                                          role='att',
                                          type='direct',
                                          help_text=f'Directing Attorney',
                                          sequence=i)
                i+=1
                SubSection.objects.create(name=f'{side} Wit {wit_num} Wit Cross',
                                          section=section,
                                          side=side,
                                          role='wit',
                                          type='cross',
                                          help_text=f'Witness (Cross)',
                                          sequence=i)
                opposing_side = 'P' if side == 'D' else 'D'
                SubSection.objects.create(name=f'{side} Wit {wit_num} Att Cross',
                                          section=section,
                                          side=opposing_side,
                                          role='att',
                                          type='cross',
                                          help_text=f'Crossing Attorney',
                                          sequence=i)
                i+=1
        closings = Section.objects.create(name='Closings', tournament=tournament)
        SubSection.objects.create(name=f'{tournament.p_choice} Closing',
                                  section=closings,
                                  side='P',
                                  role='att',
                                  type='statement',
                                  help_text=f'{tournament.p_choice} Closing',
                                  sequence=i)
        SubSection.objects.create(name=f'Defense Closing',
                                  section=closings,
                                  side='D',
                                  role='att',
                                  type='statement',
                                  help_text=f'Defense Closing',
                                  sequence=i)
    return redirect('index')


@user_passes_test(lambda u: u.is_staff)
def load_amta_witnesses(request):
    tournament = request.user.tournament
    for side, witnesses in amta_witnesses.items():
        for witness in witnesses:
            Character.objects.create(tournament=tournament, name=witness, side=side)
    return redirect('index')

def donate(request):
    return render(request, 'donate.html')

@user_passes_test(lambda u: u.is_staff)
def refresh(request):
    tournament = request.user.tournament
    teams = [team for team in Team.objects.filter(user__tournament=tournament)]
    errors = []
    for team in teams:
        team.save()
        # errors.append(f"{team} {team.total_ballots}")
        # for competitor in team.competitors.all():
        #     competitor.save()
    for team in teams:
        team.save()
    return redirect('tourney:results') #, {'errors': errors}