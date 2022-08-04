import openpyxl
import random
import string
if __name__ == '__main__':
    excel_file = '/Users/carlguo/Desktop/Judge List.xlsx'
    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb["Sheet1"]
    list = []
    n = worksheet.max_row
    m = worksheet.max_column
    for i in range(2, n + 1):
        if worksheet.cell(row=i, column=10).value == None or worksheet.cell(row=i, column=10).value.strip() == '':
            worksheet.cell(row=i, column=10).value =''.join(random.choices(string.ascii_letters + string.digits, k=4))
        if worksheet.cell(row=i, column=1).value != None:
            print(worksheet.cell(row=i, column=1).value)
            worksheet.cell(row=i, column=9).value = str(worksheet.cell(row=i, column=1).value.lower())

    wb.save("testexcel.xlsx")