import pinyin
import openpyxl
import random
import string
if __name__ == '__main__':
    excel_file = '/Users/carlguo/Desktop/JHU Tabeasy Info Template.xlsx'
    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb["Teams"]
    list = []
    n = worksheet.max_row
    m = worksheet.max_column
    for i in range(2, n + 1):
        worksheet.cell(row=i, column=17).value =''.join(random.choices(string.ascii_letters + string.digits, k=4))
        if worksheet.cell(row=i, column=2).value != None:
            worksheet.cell(row=i, column=16).value = ''.join(worksheet.cell(row=i, column=1).value.split(' '))

    worksheet = wb["Judges"]
    list = []
    n = worksheet.max_row
    m = worksheet.max_column
    for i in range(2, n + 1):
        if worksheet.cell(row=i, column=10).value == None or worksheet.cell(row=i, column=10).value.strip() == '':
            worksheet.cell(row=i, column=10).value = ''.join(random.choices(string.ascii_letters + string.digits, k=4))
        if worksheet.cell(row=i, column=1).value != None and worksheet.cell(row=i, column=2).value != None:
            worksheet.cell(row=i, column=9).value = f"{str(worksheet.cell(row=i, column=1).value.lower())}_{str(worksheet.cell(row=i, column=2).value.lower())}"


    wb.save("testexcel.xlsx")